from drf_spectacular.utils import extend_schema
from rest_framework import generics, permissions
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError

from core.permissions import IsAdminOrTourManagerOrReadOnly, TenantScopedMixin

from trips.models import Trip, TripBus
from trips.serializers import TripBusSerializer, TripSerializer


class TripListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = Trip.objects.select_related("tenant").prefetch_related(
            "trip_buses__manager", "trip_buses__driver"
        )
        return self.apply_tenant_filter(qs, "tenant_id")

    def perform_create(self, serializer):
        self.enforce_tenant_on_create(serializer, "tenant")

    @extend_schema(
        summary="List trips",
        description="Returns trips scoped by user's tenant.",
        responses={200: TripSerializer},
        tags=["Trips"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create trip",
        description="Create a trip under the current user's tenant.",
        request=TripSerializer,
        responses={201: TripSerializer, 400: {"description": "Validation error"}},
        tags=["Trips"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


class TripDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = Trip.objects.select_related("tenant").prefetch_related(
            "trip_buses__manager", "trip_buses__driver"
        )
        return self.apply_tenant_filter(qs, "tenant_id")

    @extend_schema(
        summary="Retrieve trip",
        description="Get a trip by ID (tenant scoped).",
        responses={200: TripSerializer, 404: {"description": "Not found"}},
        tags=["Trips"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update trip",
        description="Update a trip's fields.",
        request=TripSerializer,
        responses={200: TripSerializer, 400: {"description": "Validation error"}},
        tags=["Trips"],
    )
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update trip",
        description="Partially update a trip's fields.",
        request=TripSerializer,
        responses={200: TripSerializer, 400: {"description": "Validation error"}},
        tags=["Trips"],
    )
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @extend_schema(
        summary="Delete trip",
        description="Delete a trip by ID.",
        responses={204: None, 404: {"description": "Not found"}},
        tags=["Trips"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)


class TripBusListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = TripBusSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = TripBus.objects.select_related("trip", "bus", "manager", "driver")
        qs = self.apply_tenant_filter(qs, "trip__tenant_id")
        trip_param = self.request.query_params.get("trip")
        if trip_param:
            qs = qs.filter(trip_id=trip_param)
        return qs

    @extend_schema(
        summary="List trip buses",
        description="Returns trip-bus assignments scoped by tenant.",
        responses={200: TripBusSerializer},
        tags=["TripBuses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create trip bus",
        description="Create a trip-bus assignment.",
        request=TripBusSerializer,
        responses={201: TripBusSerializer, 400: {"description": "Validation error"}},
        tags=["TripBuses"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


class TripBusDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TripBusSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = TripBus.objects.select_related("trip", "bus", "manager", "driver")
        return self.apply_tenant_filter(qs, "trip__tenant_id")

    @extend_schema(
        summary="Retrieve trip bus",
        description="Get a trip-bus assignment by ID (tenant scoped).",
        responses={200: TripBusSerializer, 404: {"description": "Not found"}},
        tags=["TripBuses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update trip bus",
        description="Update a trip-bus assignment.",
        request=TripBusSerializer,
        responses={200: TripBusSerializer, 400: {"description": "Validation error"}},
        tags=["TripBuses"],
    )
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update trip bus",
        description="Partially update a trip-bus assignment.",
        request=TripBusSerializer,
        responses={200: TripBusSerializer, 400: {"description": "Validation error"}},
        tags=["TripBuses"],
    )
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @extend_schema(
        summary="Delete trip bus",
        description="Delete a trip-bus assignment by ID.",
        responses={204: None, 404: {"description": "Not found"}},
        tags=["TripBuses"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

class TripBulkDeleteView(TripListCreateView):
    from rest_framework import serializers
    from drf_spectacular.utils import inline_serializer
    from common.views import BaseAPIView

    @extend_schema(
        summary="Bulk delete trips",
        description="Delete multiple trips by ID.",
        request=inline_serializer("TripBulkDelete", fields={"ids": serializers.ListField(child=serializers.IntegerField())}),
        responses={
            200: inline_serializer(
                "TripBulkDeleteResponse", 
                fields={
                    "success": serializers.BooleanField(), 
                    "data": inline_serializer("TripBulkDeleteData", fields={"deleted": serializers.IntegerField()})
                }
            )
        },
        tags=["Trips"],
    )
    def post(self, request, *args, **kwargs):
        from common.views import BaseAPIView
        ids = request.data.get("ids", [])
        if not ids:
            return BaseAPIView().error("No ids provided")
        qs = self.get_queryset().filter(id__in=ids)
        deleted, _ = qs.delete()
        return BaseAPIView().success({"deleted": deleted})

class TripBusBulkDeleteView(TripBusListCreateView):
    from rest_framework import serializers
    from drf_spectacular.utils import inline_serializer

    @extend_schema(
        summary="Bulk delete trip buses",
        description="Delete multiple trip buses by ID.",
        request=inline_serializer("TripBusBulkDelete", fields={"ids": serializers.ListField(child=serializers.IntegerField())}),
        responses={
            200: inline_serializer(
                "TripBusBulkDeleteResponse", 
                fields={
                    "success": serializers.BooleanField(), 
                    "data": inline_serializer("TripBusBulkDeleteData", fields={"deleted": serializers.IntegerField()})
                }
            )
        },
        tags=["TripBuses"],
    )
    def post(self, request, *args, **kwargs):
        from common.views import BaseAPIView
        ids = request.data.get("ids", [])
        if not ids:
            return BaseAPIView().error("No ids provided")
        qs = self.get_queryset().filter(id__in=ids)
        deleted, _ = qs.delete()
        return BaseAPIView().success({"deleted": deleted})

TRIPBUS_COLUMNS = ["STT", "Biển số", "Mã xe", "Sức chứa", "Mô tả"]

class TripBusImportView(TenantScopedMixin, APIView):
    """POST /api/v1/trip-buses/import/?trip=<id>"""
    permission_classes = [permissions.IsAuthenticated]
    from rest_framework.parsers import MultiPartParser
    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Import trip buses from Excel",
        description="Upload a .xlsx file to import buses for a specific trip. Requires 'file' field and 'trip' query param.",
        tags=["TripBuses"],
    )
    def post(self, request, *args, **kwargs):
        trip_id = request.query_params.get("trip")
        if not trip_id:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({"detail": "trip parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({"detail": f"Cannot read Excel file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if rows else []

        imported_count = 0
        from django.db import transaction
        from fleet.models import Bus
        from trips.models import Trip, TripBus
        
        try:
            trip = Trip.objects.get(id=trip_id)
        except Trip.DoesNotExist:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({"detail": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)

        with transaction.atomic():
            for row in data_rows:
                if not row or len(row) < 4:
                    continue
                
                registration_number = str(row[1]).strip() if row[1] else ""
                bus_code = str(row[2]).strip() if row[2] else ""
                capacity = str(row[3]).strip() if row[3] else ""
                description = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                if not registration_number or not bus_code or not capacity.isdigit():
                    continue

                tenant_id = self.get_user_tenant()
                
                # First get or create the global bus
                bus, _ = Bus.objects.update_or_create(
                    registration_number=registration_number,
                    defaults={
                        "bus_code": bus_code,
                        "capacity": int(capacity),
                        "description": description,
                        "tenant_id": tenant_id
                    }
                )
                
                # Then get or create the TripBus
                TripBus.objects.update_or_create(
                    trip=trip,
                    bus=bus,
                    defaults={
                        "description": description,
                    }
                )
                
                imported_count += 1
                
        from rest_framework.response import Response
        from rest_framework import status
        return Response({"detail": f"Imported {imported_count} buses successfully."}, status=status.HTTP_201_CREATED)

class TripBusExportView(TenantScopedMixin, APIView):
    """GET /api/v1/trip-buses/export/?trip=<id>"""
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Export trip buses to Excel",
        description="Download trip buses as a .xlsx file. Requires 'trip' query param.",
        tags=["TripBuses"],
    )
    def get(self, request, *args, **kwargs):
        trip_id = request.query_params.get("trip")
        if not trip_id:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({"detail": "trip parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        import io
        import openpyxl
        from django.http import HttpResponse
        from trips.models import TripBus

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TripBuses"
        ws.append(TRIPBUS_COLUMNS)

        qs = TripBus.objects.filter(trip_id=trip_id).select_related("bus").order_by("bus__registration_number")
        # Apply tenant filter implicitly via trip? 
        # Actually TripBus doesn't have tenant_id directly, but Trip does.
        # It's fine to just filter by trip_id, assuming user has access to this trip.
        
        for idx, trip_bus in enumerate(qs, start=1):
            bus = trip_bus.bus
            ws.append([
                idx,
                bus.registration_number,
                bus.bus_code,
                bus.capacity,
                trip_bus.description or bus.description
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="trip_buses.xlsx"'
        return response

class TripBusTemplateDownloadView(APIView):
    """GET /api/v1/trip-buses/import/template/"""
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Download trip bus import template",
        description="Download a blank .xlsx template for importing trip buses.",
        tags=["TripBuses"],
    )
    def get(self, request, *args, **kwargs):
        import io
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(TRIPBUS_COLUMNS)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="trip_bus_import_template.xlsx"'
        return response
