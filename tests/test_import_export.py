import io

import openpyxl
import pytest
from django.urls import reverse
from rest_framework.test import APIClient

from accounts.models import Tenant, User
from fleet.models import Bus
from passengers.models import Passenger
from rounds.models import Round
from trips.models import Trip


@pytest.fixture
def auth_client(db):
    client = APIClient()
    tenant = Tenant.objects.create(name="Test Tenant")
    user = User.objects.create_user(
        username="testuser",
        email="test@example.com",
        password="password123",
        tenant=tenant
    )
    client.force_authenticate(user=user)
    return client


@pytest.fixture
def trip(db, auth_client):
    tenant = Tenant.objects.first()
    return Trip.objects.create(
        name="Test Trip",
        start_date="2026-05-01",
        end_date="2026-05-10",
        status="planned",
        tenant=tenant
    )


@pytest.mark.django_db
def test_passenger_import_export(auth_client, trip):
    # Test download template
    resp = auth_client.get(reverse("passenger-template"))
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Create dummy Excel file for import
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bus A"
    ws.append(["STT", "Họ và tên", "Số điện thoại", "Ghi chú"])
    ws.append([1, "John Doe", "0123456789", "Note 1"])
    ws.append([2, "Jane Doe", "0987654321", "Note 2"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "passengers.xlsx"

    # Test import (existing trip)
    resp = auth_client.post(
        reverse("passenger-import"),
        {"file": buf, "trip_id": str(trip.id)},
        format="multipart"
    )
    assert resp.status_code in [200, 201]
    assert resp.json()["trip_id"] == trip.id
    assert Passenger.objects.count() == 2

    # Test export passengers
    resp = auth_client.get(reverse("passenger-export") + f"?trip={trip.id}")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.django_db
def test_bus_import_export(auth_client):
    # Test download template
    resp = auth_client.get(reverse("bus-template"))
    assert resp.status_code == 200

    # Create dummy Excel file for import
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["STT", "Biển số", "Mã xe", "Sức chứa", "Mô tả"])
    ws.append([1, "51B-12345", "B01", 45, "Bus 1"])
    ws.append([2, "51B-54321", "B02", 30, "Bus 2"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "buses.xlsx"

    # Test import
    resp = auth_client.post(
        reverse("bus-import"),
        {"file": buf},
        format="multipart"
    )
    assert resp.status_code in [200, 201]
    assert Bus.objects.count() == 2

    # Test export
    resp = auth_client.get(reverse("bus-export"))
    assert resp.status_code == 200


@pytest.mark.django_db
def test_round_import_export(auth_client, trip):
    # Test download template
    resp = auth_client.get(reverse("round-template"))
    assert resp.status_code == 200

    # Create dummy Excel file for import
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["STT", "Tên chặng", "Địa điểm", "Thứ tự", "Ước tính", "Thực tế", "Trạng thái"])
    ws.append([1, "Round 1", "Location 1", 1, "2026-05-01 08:00:00", "", "planned"])
    ws.append([2, "Round 2", "Location 2", 2, "2026-05-01 14:00:00", "", "planned"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "rounds.xlsx"

    # Test import
    resp = auth_client.post(
        reverse("round-import") + f"?trip={trip.id}",
        {"file": buf},
        format="multipart"
    )
    assert resp.status_code in [200, 201]
    assert Round.objects.filter(trip=trip).count() == 2

    # Test export
    resp = auth_client.get(reverse("round-export") + f"?trip={trip.id}")
    assert resp.status_code == 200
