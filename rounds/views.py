import json
import logging

import paho.mqtt.publish as publish
from django.conf import settings
from django.db.models import Count, Max, Q
from drf_spectacular.utils import extend_schema
from rest_framework import generics, permissions, status

from core.permissions import (
    IsAdminOrTourManagerOrReadOnly,
    IsAdminOrTourManagerOrFleetLeadOrReadOnly,
    TenantScopedMixin,
)

from rounds.models import Round, RoundBus
from rounds.serializers import RoundBusSerializer, RoundSerializer

logger = logging.getLogger(__name__)


def publish_round_finalize_to_mqtt(payload: dict):
    try:
        if not settings.MQTT_URL:
            logger.warning("MQTT_URL not configured, skipping round finalize publish")
            return

        mqtt_url = settings.MQTT_URL.replace("wss://", "").replace("ws://", "")
        if ":" in mqtt_url:
            host, port = mqtt_url.rsplit(":", 1)
            port = int(port)
        else:
            host = mqtt_url
            port = 8883 if settings.MQTT_URL.startswith("wss") else 1883

        auth = None
        if settings.MQTT_USERNAME and settings.MQTT_PASSWORD:
            auth = {
                "username": settings.MQTT_USERNAME,
                "password": settings.MQTT_PASSWORD,
            }

        topic = f"round-finalize/{payload.get('round_bus')}"

        publish.single(
            topic,
            payload=json.dumps(payload),
            hostname=host,
            port=port,
            auth=auth,
            tls={} if settings.MQTT_URL.startswith("wss") else None,
            transport="websockets" if settings.MQTT_URL.startswith("ws") else "tcp",
        )

        logger.info("Published round finalize to MQTT topic: %s", topic)
    except Exception as exc:  # pragma: no cover - defensive log
        logger.error("Failed to publish round finalize to MQTT: %s", exc)


def sync_round_progress(round_obj: Round, *, prev_status: str | None = None) -> Round:
    """Ensure round timing/status reflect finalized buses, and start the next round when this one completes."""

    aggregates = round_obj.round_buses.aggregate(
        total=Count("id"),
        finalized=Count("id", filter=Q(finalized_at__isnull=False)),
        latest_finalized=Max("finalized_at"),
    )

    total = aggregates.get("total") or 0
    finalized = aggregates.get("finalized") or 0
    latest_finalized = aggregates.get("latest_finalized")

    updates: dict = {}

    status_changed_to_done = False

    if total and finalized == total and latest_finalized:
        # All buses closed: mark the round done and persist the real completion time.
        if round_obj.actual_time != latest_finalized:
            updates["actual_time"] = latest_finalized
        if round_obj.status != Round.Status.DONE:
            updates["status"] = Round.Status.DONE
            status_changed_to_done = prev_status != Round.Status.DONE
    else:
        # Not yet fully finalized: reflect in-progress or planned state and clear any stale time.
        next_status = Round.Status.DOING if finalized else Round.Status.PLANNED
        if round_obj.status != next_status:
            updates["status"] = next_status
        if round_obj.actual_time is not None:
            updates["actual_time"] = None

    if updates:
        Round.objects.filter(pk=round_obj.pk).update(**updates)
        for field, value in updates.items():
            setattr(round_obj, field, value)

    if status_changed_to_done:
        # If no other round is in-progress for this trip, move the next planned round into doing.
        has_doing = (
            Round.objects.filter(
                trip=round_obj.trip,
                status=Round.Status.DOING,
            )
            .exclude(pk=round_obj.pk)
            .exists()
        )

        if not has_doing:
            next_round = (
                Round.objects.filter(
                    trip=round_obj.trip,
                    sequence__gt=round_obj.sequence,
                )
                .order_by("sequence")
                .first()
            )
            if next_round and next_round.status == Round.Status.PLANNED:
                Round.objects.filter(pk=next_round.pk).update(status=Round.Status.DOING)
                next_round.status = Round.Status.DOING
            elif not next_round:
                from trips.models import Trip
                Trip.objects.filter(pk=round_obj.trip_id).update(status=Trip.Status.DONE)

    return round_obj


class RoundReorderView(generics.GenericAPIView):
    """
    POST /rounds/reorder/
    Body: [{"id": 1, "sequence": 1}, {"id": 2, "sequence": 2}, ...]
    Atomically reassigns sequence numbers to avoid unique-constraint conflicts.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        request={"application/json": {"type": "array", "items": {"type": "object"}}},
        responses={200: {"description": "Reordered successfully"}},
        tags=["Rounds"],
    )
    def post(self, request, *args, **kwargs):
        from django.db import transaction
        from rest_framework.response import Response

        items = request.data
        if not isinstance(items, list):
            return Response(
                {"detail": "Expected a list of {id, sequence} objects."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ids = [item.get("id") for item in items if item.get("id") is not None]
        if not ids:
            return Response({"detail": "Reordered successfully."}, status=status.HTTP_200_OK)

        rounds_qs = Round.objects.filter(pk__in=ids)
        round_map = {r.pk: r for r in rounds_qs}
        
        if not rounds_qs.exists():
            return Response({"detail": "Reordered successfully."}, status=status.HTTP_200_OK)
            
        trip_id = rounds_qs.first().trip_id
        
        from django.db.models import Max
        max_seq_aggr = Round.objects.filter(trip_id=trip_id).exclude(status=Round.Status.PLANNED).aggregate(Max('sequence'))
        non_planned_max_seq = max_seq_aggr['sequence__max'] or 0

        planned_items_to_update = []
        for item in items:
            rid = item.get("id")
            new_seq = item.get("sequence")
            if rid in round_map and new_seq is not None:
                r = round_map[rid]
                if r.status == Round.Status.PLANNED:
                    if new_seq <= non_planned_max_seq:
                        from rest_framework.exceptions import PermissionDenied
                        raise PermissionDenied("Không thể đưa chặng chưa đến lên trước chặng đang đi hoặc đã đi qua.")
                    planned_items_to_update.append(item)

        with transaction.atomic():
            # Step 1: push all to large temp values to avoid unique conflicts
            for item in planned_items_to_update:
                rid = item.get("id")
                Round.objects.filter(pk=rid).update(sequence=999000 + rid)

            # Step 2: apply the real new sequences
            for item in planned_items_to_update:
                rid = item.get("id")
                new_seq = item.get("sequence")
                Round.objects.filter(pk=rid).update(sequence=new_seq)

        return Response({"detail": "Reordered successfully."}, status=status.HTTP_200_OK)


class RoundListCreateView(TenantScopedMixin, generics.ListCreateAPIView):

    serializer_class = RoundSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = Round.objects.select_related("trip").prefetch_related(
            "round_buses",
            "round_buses__trip_bus",
            "round_buses__trip_bus__bus",
        )
        return self.apply_tenant_filter(qs, "trip__tenant_id")

    @extend_schema(
        summary="List rounds",
        description="Returns all rounds. If the user belongs to a tenant, results are scoped to that tenant's trips.",
        responses={200: RoundSerializer},
        tags=["Rounds"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create round",
        description="Create a round for a trip. Sequence must be unique within a trip.",
        request=RoundSerializer,
        responses={
            201: RoundSerializer,
            400: {
                "description": "Validation error",
                "example": {"sequence": ["Sequence must be unique per trip."]},
            },
        },
        tags=["Rounds"],
    )
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == status.HTTP_201_CREATED:
            response.data = {"success": True, "data": response.data}
        return response


class RoundDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = RoundSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = Round.objects.select_related("trip").prefetch_related(
            "round_buses",
            "round_buses__trip_bus",
            "round_buses__trip_bus__bus",
        )
        return self.apply_tenant_filter(qs, "trip__tenant_id")

    @extend_schema(
        summary="Retrieve round",
        description="Get a single round by ID, scoped by tenant if applicable.",
        responses={200: RoundSerializer, 404: {"description": "Round not found"}},
        tags=["Rounds"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update round",
        description="Update round fields. Sequence must remain unique within the trip.",
        request=RoundSerializer,
        responses={200: RoundSerializer, 400: {"description": "Validation error"}},
        tags=["Rounds"],
    )
    def put(self, request, *args, **kwargs):
        round_obj = self.get_object()
        if round_obj.status != "planned":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Không thể sửa chặng đã đến hoặc đang đến.")
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update round",
        description="Partially update round fields.",
        request=RoundSerializer,
        responses={200: RoundSerializer, 400: {"description": "Validation error"}},
        tags=["Rounds"],
    )
    def patch(self, request, *args, **kwargs):
        round_obj = self.get_object()
        if round_obj.status != "planned":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Không thể sửa chặng đã đến hoặc đang đến.")
        return super().patch(request, *args, **kwargs)

    @extend_schema(
        summary="Delete round",
        description="Delete a round by ID.",
        responses={204: None, 404: {"description": "Round not found"}},
        tags=["Rounds"],
    )
    def delete(self, request, *args, **kwargs):
        round_obj = self.get_object()
        if round_obj.status != "planned":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Không thể xóa chặng đã đến hoặc đang đến.")
        return super().delete(request, *args, **kwargs)


class RoundBusListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = RoundBusSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = RoundBus.objects.select_related("round", "trip_bus", "trip_bus__trip")
        return self.apply_tenant_filter(qs, "trip_bus__trip__tenant_id")

    @extend_schema(
        summary="List round-bus assignments",
        description="Returns round-bus assignments, scoped by tenant if the user belongs to one.",
        responses={200: RoundBusSerializer},
        tags=["RoundBuses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create round-bus assignment",
        description="Assign a round to a trip bus. Each round-bus pair must be unique.",
        request=RoundBusSerializer,
        responses={
            201: RoundBusSerializer,
            400: {
                "description": "Validation error",
                "example": {
                    "non_field_errors": "This round is already assigned to the bus."
                },
            },
        },
        tags=["RoundBuses"],
    )
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == status.HTTP_201_CREATED:
            response.data = {"success": True, "data": response.data}
        return response

    def perform_create(self, serializer):
        obj = serializer.save()
        sync_round_progress(obj.round, prev_status=obj.round.status)


class RoundBusDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = RoundBusSerializer
    permission_classes = [IsAdminOrTourManagerOrFleetLeadOrReadOnly]

    def get_queryset(self):
        qs = RoundBus.objects.select_related("round", "trip_bus", "trip_bus__trip")
        return self.apply_tenant_filter(qs, "trip_bus__trip__tenant_id")

    @extend_schema(
        summary="Retrieve round-bus assignment",
        description="Get a single round-bus assignment by ID, scoped by tenant if applicable.",
        responses={200: RoundBusSerializer, 404: {"description": "RoundBus not found"}},
        tags=["RoundBuses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update round-bus assignment",
        description="Update a round-bus assignment. The round and trip bus pair must stay unique.",
        request=RoundBusSerializer,
        responses={200: RoundBusSerializer, 400: {"description": "Validation error"}},
        tags=["RoundBuses"],
    )
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update round-bus assignment",
        description="Partially update a round-bus assignment.",
        request=RoundBusSerializer,
        responses={200: RoundBusSerializer, 400: {"description": "Validation error"}},
        tags=["RoundBuses"],
    )
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @extend_schema(
        summary="Delete round-bus assignment",
        description="Delete a round-bus assignment by ID.",
        responses={204: None, 404: {"description": "RoundBus not found"}},
        tags=["RoundBuses"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

    def perform_update(self, serializer):
        prev_finalized_at = getattr(serializer.instance, "finalized_at", None)
        prev_checkout_finalized_at = getattr(serializer.instance, "checkout_finalized_at", None)
        round_obj = serializer.instance.round
        prev_round_status = getattr(round_obj, "status", None)
        obj = serializer.save()
        sync_round_progress(obj.round, prev_status=prev_round_status)
        if prev_finalized_at != obj.finalized_at or prev_checkout_finalized_at != obj.checkout_finalized_at:
            payload = {
                "round_bus": obj.id,
                "round": obj.round_id,
                "trip_bus": obj.trip_bus_id,
                "trip": getattr(obj.trip_bus, "trip_id", None),
                "finalized_at": (
                    obj.finalized_at.isoformat() if obj.finalized_at else None
                ),
                "checkout_finalized_at": (
                    obj.checkout_finalized_at.isoformat() if obj.checkout_finalized_at else None
                ),
            }
            publish_round_finalize_to_mqtt(payload)
        return obj

    def perform_destroy(self, instance):
        round_obj = instance.round
        prev_round_status = getattr(round_obj, "status", None)
        super().perform_destroy(instance)
        sync_round_progress(round_obj, prev_status=prev_round_status)


ROUND_COLUMNS = ["STT", "Tên chặng", "Địa điểm", "Thời gian đến dự kiến (DD/MM/YYYY HH:MM)", "Thứ tự"]


class RoundImportView(TenantScopedMixin, generics.GenericAPIView):
    """POST /api/v1/rounds/import/?trip=<trip_id>"""

    permission_classes = [permissions.IsAuthenticated]
    
    from rest_framework.parsers import MultiPartParser
    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Import rounds from Excel",
        description="Upload a .xlsx file to import rounds. Requires 'file' field and 'trip' query parameter.",
        tags=["Rounds"],
    )
    def post(self, request, *args, **kwargs):
        from rest_framework.response import Response
        from rest_framework.views import APIView
        
        trip_id = request.query_params.get("trip")
        if not trip_id:
            return Response({"detail": "trip query param required."}, status=status.HTTP_400_BAD_REQUEST)

        from trips.models import Trip
        try:
            trip_qs = Trip.objects.all()
            trip_qs = self.apply_tenant_filter(trip_qs, "tenant_id")
            trip = trip_qs.get(pk=trip_id)
        except Trip.DoesNotExist:
            return Response({"detail": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            return Response({"detail": f"Cannot read Excel file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        action = request.data.get("action", "") # 'overwrite', 'skip', or ''
        
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if rows else []

        imported_count = 0
        from django.db import transaction
        from django.utils.dateparse import parse_datetime
        import datetime
        
        # Collect new locations to check for duplicates
        new_locations = set()
        for row in data_rows:
            if not row or len(row) < 5:
                continue
            loc = str(row[2]).strip() if row[2] else ""
            if loc:
                new_locations.add(loc)
                
        existing_rounds = Round.objects.filter(trip=trip, location__in=new_locations)
        if existing_rounds.exists() and not action:
            duplicates = [r.location for r in existing_rounds]
            return Response(
                {
                    "detail": "Có chặng trùng địa điểm",
                    "duplicates": list(set(duplicates))
                },
                status=status.HTTP_409_CONFLICT
            )

        with transaction.atomic():
            for row in data_rows:
                if not row or len(row) < 5:
                    continue
                
                name = str(row[1]).strip() if row[1] else ""
                location = str(row[2]).strip() if row[2] else ""
                estimate_time_val = row[3] if len(row) > 3 else None
                sequence_str = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                if not name or not location or not sequence_str.isdigit() or not estimate_time_val:
                    continue
                    
                sequence = int(sequence_str)
                if isinstance(estimate_time_val, datetime.datetime):
                    estimate_time = estimate_time_val
                elif isinstance(estimate_time_val, (float, int)):
                    import datetime
                    try:
                        estimate_time = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(estimate_time_val))
                    except (OverflowError, ValueError):
                        estimate_time = None
                else:
                    time_str = str(estimate_time_val).strip()
                    estimate_time = parse_datetime(time_str)
                    if not estimate_time:
                        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%y %H:%M:%S", "%d/%m/%y %H:%M"):
                            try:
                                estimate_time = datetime.datetime.strptime(time_str, fmt)
                                break
                            except ValueError:
                                pass
                if not estimate_time:
                    continue

                existing_round = Round.objects.filter(trip=trip, location=location).first()
                if existing_round:
                    if action == "skip":
                        continue
                    elif action == "overwrite":
                        existing_round.name = name
                        existing_round.sequence = sequence
                        existing_round.estimate_time = estimate_time
                        existing_round.save()
                        imported_count += 1
                else:
                    Round.objects.update_or_create(
                        trip=trip,
                        sequence=sequence,
                        defaults={
                            "name": name,
                            "location": location,
                            "estimate_time": estimate_time,
                            "status": Round.Status.PLANNED
                        }
                    )
                    imported_count += 1
                
        return Response({"detail": f"Imported {imported_count} rounds successfully."}, status=status.HTTP_201_CREATED)


class RoundExportView(TenantScopedMixin, generics.GenericAPIView):
    """GET /api/v1/rounds/export/?trip=<trip_id>"""

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Export rounds to Excel",
        description="Download all rounds for a trip as a .xlsx file.",
        tags=["Rounds"],
    )
    def get(self, request, *args, **kwargs):
        from rest_framework.response import Response
        
        trip_id = request.query_params.get("trip")
        if not trip_id:
            return Response({"detail": "trip query param required."}, status=status.HTTP_400_BAD_REQUEST)

        from trips.models import Trip
        try:
            trip_qs = Trip.objects.all()
            trip_qs = self.apply_tenant_filter(trip_qs, "tenant_id")
            trip = trip_qs.get(pk=trip_id)
        except Trip.DoesNotExist:
            return Response({"detail": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)

        import io
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rounds"
        ws.append(ROUND_COLUMNS)

        rounds = Round.objects.filter(trip=trip).order_by("sequence")
        for idx, rnd in enumerate(rounds, start=1):
            estimate_time_str = rnd.estimate_time.strftime("%d/%m/%Y %H:%M") if rnd.estimate_time else ""
            ws.append([
                idx,
                rnd.name,
                rnd.location,
                estimate_time_str,
                rnd.sequence
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="rounds_{trip.name.replace(" ", "_")}.xlsx"'
        return response


class RoundTemplateDownloadView(generics.GenericAPIView):
    """GET /api/v1/rounds/import/template/"""

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Download round import template",
        description="Download a blank .xlsx template for importing rounds.",
        tags=["Rounds"],
    )
    def get(self, request, *args, **kwargs):
        import io
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(ROUND_COLUMNS)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="round_import_template.xlsx"'
        return response

class RoundBulkDeleteView(RoundListCreateView):
    from rest_framework import serializers
    from drf_spectacular.utils import inline_serializer
    from common.views import BaseAPIView

    @extend_schema(
        summary="Bulk delete rounds",
        description="Delete multiple rounds by ID.",
        request=inline_serializer("RoundBulkDelete", fields={"ids": serializers.ListField(child=serializers.IntegerField())}),
        responses={
            200: inline_serializer(
                "RoundBulkDeleteResponse", 
                fields={
                    "success": serializers.BooleanField(), 
                    "data": inline_serializer("RoundBulkDeleteData", fields={"deleted": serializers.IntegerField()})
                }
            )
        },
        tags=["Rounds"],
    )
    def post(self, request, *args, **kwargs):
        from common.views import BaseAPIView
        ids = request.data.get("ids", [])
        if not ids:
            return BaseAPIView().error("No ids provided")
        qs = self.get_queryset().filter(id__in=ids)
        
        for r in qs:
            if r.status != "planned":
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Không thể xóa chặng đã đến hoặc đang đến.")
                
        deleted, _ = qs.delete()
        return BaseAPIView().success({"deleted": deleted})
