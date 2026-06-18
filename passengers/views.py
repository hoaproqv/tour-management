import io
import json
import logging

import paho.mqtt.publish as publish
from django.conf import settings
from django.db import transaction
from django.db.models import Prefetch
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from rest_framework import generics, permissions, status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import (
    IsAdminOrTourManagerOrFleetLeadOrReadOnly,
    IsAdminOrTourManagerOrReadOnly,
    TenantScopedMixin,
)
from passengers.models import (
    ImportedBus,
    Passenger,
    PassengerBusAssignment,
    PassengerTransfer,
)
from passengers.serializers import (
    ImportedBusSerializer,
    PassengerAssignmentSerializer,
    PassengerSerializer,
    PassengerTransferSerializer,
)
from trips.models import Trip, TripBus

logger = logging.getLogger(__name__)

PASSENGER_COLUMNS = ["STT", "Họ và tên", "Số điện thoại", "Thông tin thêm", "Ghi chú"]


def publish_transfer_to_mqtt(payload: dict):
    try:
        if not settings.MQTT_URL:
            logger.warning("MQTT_URL not configured, skipping transfer publish")
            return

        mqtt_url = settings.MQTT_URL.replace("wss://", "").replace("ws://", "")
        if ":" in mqtt_url:
            host, port = mqtt_url.rsplit(":", 1)
            port = int(port)
        else:
            host = mqtt_url
            port = 8883 if settings.MQTT_URL.startswith("wss") else 1883

        auth = None
        if settings.MQTT_USERNAME and settings.MQTT_PASSWORD:
            auth = {
                "username": settings.MQTT_USERNAME,
                "password": settings.MQTT_PASSWORD,
            }

        topic = f"passenger-transfer/{payload.get('passenger')}"

        publish.single(
            topic,
            payload=json.dumps(payload),
            hostname=host,
            port=port,
            auth=auth,
            tls={} if settings.MQTT_URL.startswith("wss") else None,
            transport="websockets" if settings.MQTT_URL.startswith("ws") else "tcp",
        )

        logger.info("Published passenger transfer to MQTT topic: %s", topic)
    except Exception as exc:  # pragma: no cover - defensive log
        logger.error("Failed to publish transfer to MQTT: %s", exc)


class PassengerListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = PassengerSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        trip_id = self.request.query_params.get("trip")
        assignment_qs = PassengerBusAssignment.objects.select_related(
            "trip_bus",
            "trip",
        )
        if trip_id:
            assignment_qs = assignment_qs.filter(trip_id=trip_id)

        qs = Passenger.objects.prefetch_related(
            Prefetch("bus_assignments", queryset=assignment_qs),
            Prefetch("bus_assignments", queryset=PassengerBusAssignment.objects.select_related("trip"), to_attr="all_assignments"),
        )
        qs = self.apply_tenant_filter(qs, "tenant_id")
        if trip_id:
            qs = qs.filter(bus_assignments__trip_id=trip_id)
        return qs.distinct()

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx.update(
            {
                "trip_id": self.request.query_params.get("trip"),
                "tenant": getattr(self.request.user, "tenant", None),
            }
        )
        return ctx

    def perform_create(self, serializer):
        self.enforce_tenant_on_create(serializer, "tenant")

    @extend_schema(
        summary="List passengers",
        description="Returns passengers, scoped to user's tenant; optionally filter by trip assignments.",
        responses={200: PassengerSerializer},
        tags=["Passengers"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create passenger",
        description="Create a passenger (tenant scoped).",
        request=PassengerSerializer,
        responses={201: PassengerSerializer, 400: {"description": "Validation error"}},
        tags=["Passengers"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


class PassengerDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PassengerSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        trip_id = self.request.query_params.get("trip")
        assignment_qs = PassengerBusAssignment.objects.select_related(
            "trip_bus",
            "trip",
        )
        if trip_id:
            assignment_qs = assignment_qs.filter(trip_id=trip_id)

        qs = Passenger.objects.prefetch_related(
            Prefetch("bus_assignments", queryset=assignment_qs),
            Prefetch("bus_assignments", queryset=PassengerBusAssignment.objects.select_related("trip"), to_attr="all_assignments"),
        )
        qs = self.apply_tenant_filter(qs, "tenant_id")
        if trip_id:
            qs = qs.filter(bus_assignments__trip_id=trip_id)
        return qs.distinct()

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx.update(
            {
                "trip_id": self.request.query_params.get("trip"),
                "tenant": getattr(self.request.user, "tenant", None),
            }
        )
        return ctx

    @extend_schema(
        summary="Retrieve passenger",
        description="Get a passenger by ID (tenant scoped).",
        responses={200: PassengerSerializer, 404: {"description": "Not found"}},
        tags=["Passengers"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update passenger",
        description="Update passenger fields.",
        request=PassengerSerializer,
        responses={200: PassengerSerializer, 400: {"description": "Validation error"}},
        tags=["Passengers"],
    )
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update passenger",
        description="Partially update passenger fields.",
        request=PassengerSerializer,
        responses={200: PassengerSerializer, 400: {"description": "Validation error"}},
        tags=["Passengers"],
    )
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @extend_schema(
        summary="Delete passenger",
        description="Delete a passenger by ID.",
        responses={204: None, 404: {"description": "Not found"}},
        tags=["Passengers"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)


class PassengerTransferListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = PassengerTransferSerializer
    permission_classes = [IsAdminOrTourManagerOrFleetLeadOrReadOnly]
    pagination_class = None

    def get_queryset(self):
        qs = PassengerTransfer.objects.select_related(
            "passenger",
            "from_trip_bus",
            "to_trip_bus",
            "trip",
        )
        qs = self.apply_tenant_filter(qs, "passenger__tenant_id")
        trip = self.request.query_params.get("trip")
        if trip:
            qs = qs.filter(trip_id=trip)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        payload = {
            "id": instance.id,
            "passenger": instance.passenger_id,
            "from_trip_bus": instance.from_trip_bus_id,
            "to_trip_bus": instance.to_trip_bus_id,
            "trip": instance.trip_id,
            "deleted": False,
        }
        publish_transfer_to_mqtt(payload)

    @extend_schema(
        summary="List passenger transfers",
        description="List transfers; scoped by tenant and optional trip.",
        responses={200: PassengerTransferSerializer},
        tags=["PassengerTransfers"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create or update passenger transfer",
        description="Create or replace the active transfer for a passenger.",
        request=PassengerTransferSerializer,
        responses={201: PassengerTransferSerializer, 200: PassengerTransferSerializer},
        tags=["PassengerTransfers"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


class PassengerTransferDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PassengerTransferSerializer
    permission_classes = [IsAdminOrTourManagerOrFleetLeadOrReadOnly]

    def get_queryset(self):
        qs = PassengerTransfer.objects.select_related(
            "passenger",
            "from_trip_bus",
            "to_trip_bus",
            "trip",
        )
        return self.apply_tenant_filter(qs, "passenger__tenant_id")

    def perform_update(self, serializer):
        instance = serializer.save()
        payload = {
            "id": instance.id,
            "passenger": instance.passenger_id,
            "from_trip_bus": instance.from_trip_bus_id,
            "to_trip_bus": instance.to_trip_bus_id,
            "trip": instance.trip_id,
            "deleted": False,
        }
        publish_transfer_to_mqtt(payload)

    def perform_destroy(self, instance):
        payload = {
            "id": instance.id,
            "passenger": instance.passenger_id,
            "from_trip_bus": instance.from_trip_bus_id,
            "to_trip_bus": instance.to_trip_bus_id,
            "trip": instance.trip_id,
            "deleted": True,
        }
        super().perform_destroy(instance)
        publish_transfer_to_mqtt(payload)

    @extend_schema(
        summary="Retrieve passenger transfer",
        responses={200: PassengerTransferSerializer, 404: {"description": "Not found"}},
        tags=["PassengerTransfers"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update passenger transfer",
        request=PassengerTransferSerializer,
        responses={200: PassengerTransferSerializer},
        tags=["PassengerTransfers"],
    )
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Delete passenger transfer",
        responses={204: None},
        tags=["PassengerTransfers"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)


class PassengerAssignmentListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = PassengerAssignmentSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]
    pagination_class = None

    def get_queryset(self):
        qs = PassengerBusAssignment.objects.select_related(
            "passenger",
            "trip_bus",
            "trip",
        )
        qs = self.apply_tenant_filter(qs, "trip__tenant_id")
        trip = self.request.query_params.get("trip")
        if trip:
            qs = qs.filter(trip_id=trip)
        return qs

    @extend_schema(
        summary="List passenger-bus assignments",
        description="List passenger assignments by trip (tenant scoped).",
        responses={200: PassengerAssignmentSerializer},
        tags=["PassengerAssignments"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Upsert passenger-bus assignment",
        description="Create or replace an assignment for a passenger.",
        request=PassengerAssignmentSerializer,
        responses={
            201: PassengerAssignmentSerializer,
            200: PassengerAssignmentSerializer,
        },
        tags=["PassengerAssignments"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


class PassengerAssignmentDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PassengerAssignmentSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = PassengerBusAssignment.objects.select_related(
            "passenger",
            "trip_bus",
            "trip",
        )
        return self.apply_tenant_filter(qs, "trip__tenant_id")

    @extend_schema(
        summary="Retrieve passenger-bus assignment",
        responses={
            200: PassengerAssignmentSerializer,
            404: {"description": "Not found"},
        },
        tags=["PassengerAssignments"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Delete passenger-bus assignment",
        responses={204: None},
        tags=["PassengerAssignments"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Import / Export
# ---------------------------------------------------------------------------

class PassengerImportCheckView(TenantScopedMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Check passengers import",
        tags=["Passengers"],
    )
    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        tenant_id = self.get_user_tenant()
        if not tenant_id:
            return Response({"detail": "Cần có tenant để check."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as e:
            return Response({"detail": f"Error reading Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        valid_passengers = []
        conflicts = []

        for sheet_name in wb.sheetnames:
            if sheet_name == "Quản lý xe":
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            data_rows = rows[1:] if rows else []

            for row in data_rows:
                if not row or not any(row):
                    continue
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                extra_info = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                note = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                if not name:
                    continue

                row_data = {
                    "name": name,
                    "phone": phone if phone.lower() != 'none' else "",
                    "extra_info": extra_info if extra_info.lower() != 'none' else "",
                    "note": note if note.lower() != 'none' else "",
                    "sheet_name": sheet_name
                }

                conflict_passenger = None
                if row_data["phone"]:
                    conflict_passenger = Passenger.objects.filter(tenant_id=tenant_id, phone=row_data["phone"]).first()

                if conflict_passenger and conflict_passenger.name != name:
                    conflicts.append({
                        "imported": row_data,
                        "existing": {
                            "id": conflict_passenger.id,
                            "name": conflict_passenger.name,
                            "phone": conflict_passenger.phone,
                        }
                    })
                else:
                    valid_passengers.append(row_data)

        return Response({
            "valid_passengers": valid_passengers,
            "conflicts": conflicts
        })


class PassengerImportView(TenantScopedMixin, APIView):
    """POST /api/v1/passengers/import/

    Multipart form with:
      - file: .xlsx file
      - trip_id (optional): existing trip PK
      - trip_name, trip_start_date, trip_end_date (optional): create new trip
    """

    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Import passengers from Excel",
        description=(
            "Upload a .xlsx file where each sheet is a bus group. "
            "Provide trip_id to use an existing trip, or trip_name + dates to create a new one. "
            "Returns created ImportedBus records that must be mapped to real Bus entities."
        ),
        tags=["Passengers"],
    )
    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        import json
        conflict_resolutions_str = request.data.get("conflict_resolutions", "{}")
        try:
            conflict_resolutions = json.loads(conflict_resolutions_str)
        except:
            conflict_resolutions = {}

        # --- Resolve / create trip ---
        trip_id = request.data.get("trip_id")
        trip = None
        if trip_id:
            try:
                trip = Trip.objects.get(pk=trip_id)
            except Trip.DoesNotExist:
                return Response({"detail": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            trip_name = request.data.get("trip_name", "").strip()
            trip_start = request.data.get("trip_start_date", "")
            trip_end = request.data.get("trip_end_date", "")
            if not trip_name or not trip_start or not trip_end:
                return Response(
                    {"detail": "Provide trip_id or trip_name + trip_start_date + trip_end_date."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            tenant_id = self.get_user_tenant()
            if not tenant_id:
                return Response(
                    {"detail": "Cần có tenant để tạo tour mới. Vui lòng chọn tour có sẵn."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            trip = Trip.objects.create(
                tenant_id=tenant_id,
                name=trip_name,
                start_date=trip_start,
                end_date=trip_end,
                status=Trip.Status.PLANNED,
            )

        # Always derive tenant from the trip — works for any user (incl. superusers)
        tenant_id = trip.tenant_id
        if not tenant_id:
            return Response(
                {"detail": "Tour này chưa được gán công ty. Vui lòng cập nhật lại tour."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- Parse Excel ---
        try:
            import openpyxl  # lazy import
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        except ModuleNotFoundError:
            return Response({"detail": "openpyxl not installed on server. Run: pip install openpyxl"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as exc:
            return Response({"detail": f"Cannot read Excel file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        result_buses = []

        with transaction.atomic():
            for seq, sheet_name in enumerate(wb.sheetnames, start=1):
                if sheet_name == "Quản lý xe":
                    continue

                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Skip header row (first row)
                data_rows = rows[1:] if rows else []

                # Filter out truly blank rows first (rows must have a name in col 1)
                non_empty_rows = [
                    r for r in data_rows
                    if r and len(r) > 1 and r[1] and str(r[1]).strip()
                ]

                # Skip sheet entirely if no passengers with a name
                if not non_empty_rows:
                    continue

                # Get or create ImportedBus for this sheet
                imported_bus, _ = ImportedBus.objects.get_or_create(
                    trip=trip,
                    sheet_name=sheet_name,
                    defaults={"sequence": seq},
                )
                imported_bus.sequence = seq
                imported_bus.save(update_fields=["sequence"])

                passenger_count = 0
                for row in non_empty_rows:
                    # Columns: STT | Họ và tên | Số điện thoại | Ghi chú
                    if not row or not any(row):
                        continue
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    phone = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    phone = phone if phone.lower() != 'none' else ""
                    extra_info = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    extra_info = extra_info if extra_info.lower() != 'none' else ""
                    note = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    note = note if note.lower() != 'none' else ""

                    if not name:
                        continue

                    passenger = None
                    if phone and tenant_id:
                        passenger = Passenger.objects.filter(
                            tenant_id=tenant_id, phone=phone, bus_assignments__trip=trip
                        ).first()
                    if not passenger and tenant_id:
                        passenger = Passenger.objects.filter(
                            tenant_id=tenant_id, name=name, bus_assignments__trip=trip
                        ).first()

                    if not passenger:
                        passenger = Passenger.objects.create(
                            tenant_id=tenant_id,
                            name=name,
                            phone=phone,
                            extra_info=extra_info,
                            note=note,
                        )
                    else:
                        # Apply conflict resolution
                        res = conflict_resolutions.get(phone)
                        if res == "update" and passenger.name != name:
                            passenger.name = name
                            passenger.save(update_fields=["name"])

                        # Update note/extra_info if provided and previously blank
                        fields_to_update = []
                        if note and not passenger.note:
                            passenger.note = note
                            fields_to_update.append("note")
                        if extra_info and not passenger.extra_info:
                            passenger.extra_info = extra_info
                            fields_to_update.append("extra_info")

                        if fields_to_update:
                            passenger.save(update_fields=fields_to_update)

                    # Upsert assignment to imported_bus (draft)
                    PassengerBusAssignment.objects.update_or_create(
                        passenger=passenger,
                        trip=trip,
                        defaults={"trip_bus": None, "imported_bus": imported_bus},
                    )
                    passenger_count += 1

                result_buses.append({
                    "id": imported_bus.id,
                    "sheet_name": sheet_name,
                    "sequence": seq,
                    "passenger_count": passenger_count,
                    "is_mapped": imported_bus.mapped_bus_id is not None,
                })

        return Response(
            {
                "trip_id": trip.id,
                "trip_name": trip.name,
                "imported_buses": result_buses,
            },
            status=status.HTTP_201_CREATED,
        )


class PassengerExportView(TenantScopedMixin, APIView):
    """GET /api/v1/passengers/export/?trip=<id>

    Returns a .xlsx file where each sheet is a TripBus (or ImportedBus if unmapped).
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Export passengers to Excel",
        description="Download passengers grouped by bus as sheets in a .xlsx file.",
        tags=["Passengers"],
    )
    def get(self, request, *args, **kwargs):
        trip_id = request.query_params.get("trip")

        if not trip_id:
            return Response({"detail": "trip query param required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            trip_qs = Trip.objects.all()
            trip_qs = self.apply_tenant_filter(trip_qs, "tenant_id")
            trip = trip_qs.get(pk=trip_id)
        except Trip.DoesNotExist:
            return Response({"detail": "Trip not found."}, status=status.HTTP_404_NOT_FOUND)

        import openpyxl  # lazy import
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # --- Sheets from TripBus (mapped buses) ---
        trip_buses = TripBus.objects.filter(trip=trip).select_related("bus").order_by("bus__registration_number")

        for tb in trip_buses:
            sheet_title = (
                getattr(tb.bus, "registration_number", None)
                or getattr(tb.bus, "bus_code", None)
                or f"Xe {tb.id}"
            )
            # Excel sheet names must be <= 31 chars
            sheet_title = sheet_title[:31]
            ws = wb.create_sheet(title=sheet_title)
            ws.append(PASSENGER_COLUMNS)
            ws.column_dimensions['C'].number_format = '@'
            assignments = (
                PassengerBusAssignment.objects.filter(trip=trip, trip_bus=tb)
                .select_related("passenger")
                .order_by("passenger__name")
            )
            for idx, a in enumerate(assignments, start=1):
                p = a.passenger
                row_idx = idx + 1
                ws.append([idx, p.name, p.phone, p.extra_info, p.note])
                ws.cell(row=row_idx, column=3).number_format = '@'

        # --- Sheets from ImportedBus (unmapped draft buses) ---
        unmapped_buses = ImportedBus.objects.filter(
            trip=trip, mapped_bus__isnull=True
        ).order_by("sequence")

        for ib in unmapped_buses:
            sheet_title = ib.sheet_name[:31]
            ws = wb.create_sheet(title=sheet_title)
            ws.append(PASSENGER_COLUMNS)
            ws.column_dimensions['C'].number_format = '@'

            assignments = (
                PassengerBusAssignment.objects.filter(trip=trip, imported_bus=ib)
                .select_related("passenger")
                .order_by("passenger__name")
            )
            for idx, a in enumerate(assignments, start=1):
                p = a.passenger
                row_idx = idx + 1
                ws.append([idx, p.name, p.phone, p.extra_info, p.note])
                ws.cell(row=row_idx, column=3).number_format = '@'

        # Fallback sheet for passengers with no bus assignment
        unassigned = (
            PassengerBusAssignment.objects.filter(
                trip=trip, trip_bus__isnull=True, imported_bus__isnull=True
            )
            .select_related("passenger")
            .order_by("passenger__name")
        )
        if unassigned.exists():
            ws = wb.create_sheet(title="Chưa gán xe")
            ws.append(PASSENGER_COLUMNS)
            ws.column_dimensions['C'].number_format = '@'
            for idx, a in enumerate(unassigned, start=1):
                p = a.passenger
                row_idx = idx + 1
                ws.append([idx, p.name, p.phone, p.extra_info, p.note])
                ws.cell(row=row_idx, column=3).number_format = '@'

        if not wb.sheetnames:
            wb.create_sheet(title="Sheet1")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"passengers_{trip.name.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PassengerTemplateDownloadView(APIView):
    """GET /api/v1/passengers/import/template/"""

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Download passenger import template",
        description="Download a blank .xlsx template for importing passengers.",
        tags=["Passengers"],
    )
    def get(self, request, *args, **kwargs):
        import io

        import openpyxl

        wb = openpyxl.Workbook()

        ws_xe1 = wb.active
        ws_xe1.title = "Xe 1"
        ws_xe1.append(PASSENGER_COLUMNS)
        ws_xe1.column_dimensions['C'].number_format = '@'
        for row in range(2, 500):
            ws_xe1.cell(row=row, column=3).number_format = '@'

        ws_xe2 = wb.create_sheet(title="Xe 2")
        ws_xe2.append(PASSENGER_COLUMNS)
        ws_xe2.column_dimensions['C'].number_format = '@'
        for row in range(2, 500):
            ws_xe2.cell(row=row, column=3).number_format = '@'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="passenger_import_template.xlsx"'
        return response


class ImportedBusListView(TenantScopedMixin, generics.ListAPIView):
    """GET /api/v1/imported-buses/?trip=<id>"""

    serializer_class = ImportedBusSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = ImportedBus.objects.select_related("trip", "mapped_bus", "mapped_trip_bus")
        qs = self.apply_tenant_filter(qs, "trip__tenant_id")
        trip_id = self.request.query_params.get("trip")
        if trip_id:
            qs = qs.filter(trip_id=trip_id)
        return qs.order_by("sequence")

    @extend_schema(
        summary="List imported buses",
        description="List draft buses created from Excel import. Filter by trip.",
        tags=["ImportedBuses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class ImportedBusMapView(TenantScopedMixin, APIView):
    """PATCH /api/v1/imported-buses/<pk>/map/

    Body: { trip_bus_id }
    Maps an imported bus sheet to an existing TripBus and updates all passenger assignments.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Map imported bus to a TripBus",
        description=(
            "Assign an existing TripBus (by trip_bus_id) to a draft ImportedBus. "
            "This moves all passenger assignments to it."
        ),
        tags=["ImportedBuses"],
    )
    def patch(self, request, pk, *args, **kwargs):
        try:
            qs = ImportedBus.objects.select_related("trip")
            qs = self.apply_tenant_filter(qs, "trip__tenant_id")
            imported_bus = qs.get(pk=pk)
        except ImportedBus.DoesNotExist:
            return Response({"detail": "ImportedBus not found."}, status=status.HTTP_404_NOT_FOUND)

        trip_bus_id = request.data.get("trip_bus_id")

        if not trip_bus_id:
            return Response(
                {"detail": "trip_bus_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            trip_bus = TripBus.objects.get(pk=trip_bus_id, trip=imported_bus.trip)
        except TripBus.DoesNotExist:
            return Response({"detail": "TripBus not found in this trip."}, status=status.HTTP_404_NOT_FOUND)

        with transaction.atomic():
            # Update imported_bus record
            imported_bus.mapped_bus = trip_bus.bus
            imported_bus.mapped_trip_bus = trip_bus
            imported_bus.save(update_fields=["mapped_bus", "mapped_trip_bus"])

            # Move all passenger assignments from draft to real trip_bus
            PassengerBusAssignment.objects.filter(
                imported_bus=imported_bus
            ).update(trip_bus=trip_bus)

        serializer = ImportedBusSerializer(imported_bus)
        return Response(serializer.data, status=status.HTTP_200_OK)


class PassengerBulkDeleteView(PassengerListCreateView):
    from drf_spectacular.utils import inline_serializer
    from rest_framework import serializers

    from common.views import BaseAPIView

    @extend_schema(
        summary="Bulk delete passengers",
        description="Delete multiple passengers by ID.",
        request=inline_serializer("PassengerBulkDelete", fields={"ids": serializers.ListField(child=serializers.IntegerField())}),
        responses={
            200: inline_serializer(
                "PassengerBulkDeleteResponse",
                fields={
                    "success": serializers.BooleanField(),
                    "data": inline_serializer("PassengerBulkDeleteData", fields={"deleted": serializers.IntegerField()})
                }
            )
        },
        tags=["Passengers"],
    )
    def post(self, request, *args, **kwargs):
        from common.views import BaseAPIView
        ids = request.data.get("ids", [])
        if not ids:
            return BaseAPIView().error("No ids provided")
        qs = self.get_queryset().filter(id__in=ids)
        deleted, _ = qs.delete()
        return BaseAPIView().success({"deleted": deleted})
