from drf_spectacular.utils import extend_schema
from rest_framework import filters, generics, permissions
from rest_framework.views import APIView

from core.permissions import IsAdminOrTourManagerOrReadOnly, TenantScopedMixin
from fleet.models import Bus
from fleet.serializers import BusSerializer


class BusListCreateView(TenantScopedMixin, generics.ListCreateAPIView):
    serializer_class = BusSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ["registration_number", "bus_code", "description"]

    def get_queryset(self):
        qs = Bus.objects.all().order_by("registration_number")
        return self.apply_tenant_filter(qs, "tenant_id")

    @extend_schema(
        summary="List buses",
        description="Return all buses ordered by registration number.",
        responses={200: BusSerializer},
        tags=["Buses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Create bus",
        description="Create a new bus record.",
        request=BusSerializer,
        responses={201: BusSerializer, 400: {"description": "Validation error"}},
        tags=["Buses"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)

    def perform_create(self, serializer):
        self.enforce_tenant_on_create(serializer, "tenant")


class BusDetailView(TenantScopedMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = BusSerializer
    permission_classes = [IsAdminOrTourManagerOrReadOnly]

    def get_queryset(self):
        qs = Bus.objects.all().order_by("registration_number")
        return self.apply_tenant_filter(qs, "tenant_id")

    @extend_schema(
        summary="Retrieve bus",
        description="Get a bus by ID.",
        responses={200: BusSerializer, 404: {"description": "Not found"}},
        tags=["Buses"],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @extend_schema(
        summary="Update bus",
        description="Update all fields of a bus.",
        request=BusSerializer,
        responses={200: BusSerializer, 400: {"description": "Validation error"}},
        tags=["Buses"],
    )
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update bus",
        description="Partially update fields of a bus.",
        request=BusSerializer,
        responses={200: BusSerializer, 400: {"description": "Validation error"}},
        tags=["Buses"],
    )
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @extend_schema(
        summary="Delete bus",
        description="Delete a bus by ID.",
        responses={204: None, 404: {"description": "Not found"}},
        tags=["Buses"],
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)


BUS_COLUMNS = ["STT", "Biển số", "Mã xe", "Sức chứa", "Mô tả"]


class BusImportView(TenantScopedMixin, APIView):
    """POST /api/v1/buses/import/"""

    permission_classes = [permissions.IsAuthenticated]

    from rest_framework.parsers import MultiPartParser
    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Import buses from Excel",
        description="Upload a .xlsx file to import buses. Requires 'file' field.",
        tags=["Buses"],
    )
    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            from rest_framework import status
            from rest_framework.response import Response
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            from rest_framework import status
            from rest_framework.response import Response
            return Response({"detail": f"Cannot read Excel file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if rows else []

        imported_count = 0
        from django.db import transaction
        with transaction.atomic():
            for row in data_rows:
                if not row or len(row) < 4:
                    continue

                registration_number = str(row[1]).strip() if row[1] else ""
                bus_code = str(row[2]).strip() if row[2] else ""
                capacity = str(row[3]).strip() if row[3] else ""
                description = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                if not registration_number or not bus_code or not capacity.isdigit():
                    continue

                tenant_id = self.get_user_tenant()

                Bus.objects.update_or_create(
                    registration_number=registration_number,
                    defaults={
                        "bus_code": bus_code,
                        "capacity": int(capacity),
                        "description": description,
                        "tenant_id": tenant_id
                    }
                )
                imported_count += 1
        from rest_framework import status
        from rest_framework.response import Response
        return Response({"detail": f"Imported {imported_count} buses successfully."}, status=status.HTTP_201_CREATED)


class BusExportView(TenantScopedMixin, APIView):
    """GET /api/v1/buses/export/"""

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Export buses to Excel",
        description="Download all buses as a .xlsx file.",
        tags=["Buses"],
    )
    def get(self, request, *args, **kwargs):
        import io

        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Buses"
        ws.append(BUS_COLUMNS)

        qs = Bus.objects.all().order_by("registration_number")
        buses = self.apply_tenant_filter(qs, "tenant_id")
        for idx, bus in enumerate(buses, start=1):
            ws.append([
                idx,
                bus.registration_number,
                bus.bus_code,
                bus.capacity,
                bus.description
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="buses.xlsx"'
        return response


class BusTemplateDownloadView(APIView):
    """GET /api/v1/buses/import/template/"""

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Download bus import template",
        description="Download a blank .xlsx template for importing buses.",
        tags=["Buses"],
    )
    def get(self, request, *args, **kwargs):
        import io

        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(BUS_COLUMNS)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="bus_import_template.xlsx"'
        return response


class BusBulkDeleteView(BusListCreateView):
    from drf_spectacular.utils import inline_serializer
    from rest_framework import serializers

    from common.views import BaseAPIView

    @extend_schema(
        summary="Bulk delete buses",
        description="Delete multiple buses by ID.",
        request=inline_serializer("BusBulkDelete", fields={"ids": serializers.ListField(child=serializers.IntegerField())}),
        responses={
            200: inline_serializer(
                "BusBulkDeleteResponse",
                fields={
                    "success": serializers.BooleanField(),
                    "data": inline_serializer("BusBulkDeleteData", fields={"deleted": serializers.IntegerField()})
                }
            )
        },
        tags=["Buses"],
    )
    def post(self, request, *args, **kwargs):
        from common.views import BaseAPIView
        ids = request.data.get("ids", [])
        if not ids:
            return BaseAPIView().error("No ids provided")
        qs = self.get_queryset().filter(id__in=ids)
        deleted, _ = qs.delete()
        return BaseAPIView().success({"deleted": deleted})
